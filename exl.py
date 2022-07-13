# -*- coding: utf-8 -*-

import os
import time
import threading
import openpyxl
import gspread

class ExcelType():
	def __init__(self, path, name, date, time):
		self.path = path
		self.name = name
		self.date = date
		self.time = time

class Excel():
	def __init__(self):
		self.db = []
		
	def start(self):
		threading.Thread(target = self.loop, daemon = True, args=()).start()
			
	def write(self, path, name, date, time):
		print('[SIGN] %s, %s, %s' % (date, time, name))
		self.db.append(ExcelType(path, name, date, time))

	def loop(self):	
		while True:
			def write(type):
				str = '%s, %s, %s' % (type.date, type.time, type.name)
				
				ret = 0
				try:
					xlsx = type.path + 'log.xlsx'
					if not os.path.isfile(xlsx):
						wb = openpyxl.Workbook()
					else:
						wb = openpyxl.load_workbook(xlsx)
						
					sheet_names = wb.get_sheet_names()
					ws = wb.get_sheet_by_name(sheet_names[0])
				
					def findDate(ws, date):
						for i in range(ws.max_column):
							col = i + 1
							cell = ws.cell(column = col, row = 1).value
							if cell == date:
								return col
						return ws.max_column + 1
					
					def findName(ws, name):
						for i in range(ws.max_row):
							row = i + 1
							cell = ws.cell(column = 1, row = row).value
							if cell == name:
								return row
						return ws.max_row + 1
					
					col, row = findDate(ws, type.date), findName(ws, type.name)
					if ws.cell(row = row, column = 1).value == None:
						ws.cell(row = row, column = 1).value = type.name
					if ws.cell(row = 1, column = col).value == None:
						ws.cell(row = 1, column = col).value = type.date
					if ws.cell(row = row, column = col).value == None:
						ws.cell(row = row, column = col).value = type.time
					wb.save(xlsx)
				except:
					print('[ERROR] write xlsx failed. please try to "close xlsx".')
					str = str + ' [xlsx error]'
					ret = 1
				
				try:
					txt = type.path + 'log.txt'
					f = open(txt, 'a')
					f.write(str + '\n')
				except:
					print('[ERROR] write txt failed')
					ret = 3
				return ret
				
			if len(self.db) > 0:
				for data in self.db:
					sleep = 0
					if not write(data):
						self.db.remove(data)
					else:
						sleep = 1
					time.sleep(sleep)
			else:
				time.sleep(1)